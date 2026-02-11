"""Excel and CSV output writer for generated reconciliation data."""

import csv
from pathlib import Path
from typing import List
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rich.progress import Progress, TaskID

from .scenarios.base import ReconciliationScenario
from .models import GenerationStats


class ExcelWriter:
    """Writes reconciliation data to a timestamped subfolder.

    Produces:
      - <subfolder>/<name>.xlsx   (two data sheets + stats sheet)
      - <subfolder>/<dataset1>.csv
      - <subfolder>/<dataset2>.csv
    """
    
    # Styling constants
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, output_dir: str, scenario: ReconciliationScenario):
        """Initialise writer.

        Parameters
        ----------
        output_dir : str
            Path to the timestamped output **folder** (will be created).
        scenario : ReconciliationScenario
            The active scenario (provides column schemas & names).
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        self.scenario = scenario
        self.workbook = Workbook()
        
        # Create sheets
        self.sheet1 = self.workbook.active
        self.sheet1.title = scenario.dataset1_name[:31]  # Excel limit
        
        self.sheet2 = self.workbook.create_sheet(title=scenario.dataset2_name[:31])
        
        # Write headers
        self._write_headers()
        
        # Track row positions
        self.sheet1_row = 2
        self.sheet2_row = 2

        # Accumulate records for CSV export
        self._source_records: List[dict] = []
        self._target_records: List[dict] = []
    
    def _write_headers(self):
        """Write and style headers for both sheets."""
        # Sheet 1 headers
        columns1 = self.scenario.get_dataset1_columns()
        for col_idx, col_name in enumerate(columns1, 1):
            cell = self.sheet1.cell(row=1, column=col_idx, value=col_name)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER
        
        # Sheet 2 headers
        columns2 = self.scenario.get_dataset2_columns()
        for col_idx, col_name in enumerate(columns2, 1):
            cell = self.sheet2.cell(row=1, column=col_idx, value=col_name)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER
    
    @staticmethod
    def _format_value(value):
        """Format a value for Excel (converts dates to ISO strings to avoid
        openpyxl serialisation issues that produce corrupt .xlsx files)."""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, date):
            return value.isoformat()
        return value
    
    def write_chunk(
        self, 
        source_records: List[dict], 
        target_records: List[dict]
    ):
        """Write a chunk of records to both sheets and accumulate for CSV."""
        columns1 = self.scenario.get_dataset1_columns()
        columns2 = self.scenario.get_dataset2_columns()
        
        # Write source records to sheet 1
        for record in source_records:
            for col_idx, col_name in enumerate(columns1, 1):
                value = self._format_value(record.get(col_name))
                cell = self.sheet1.cell(row=self.sheet1_row, column=col_idx, value=value)
                cell.border = self.BORDER
            self.sheet1_row += 1
        
        # Write target records to sheet 2
        for record in target_records:
            for col_idx, col_name in enumerate(columns2, 1):
                value = self._format_value(record.get(col_name))
                cell = self.sheet2.cell(row=self.sheet2_row, column=col_idx, value=value)
                cell.border = self.BORDER
            self.sheet2_row += 1

        # Accumulate for CSV
        self._source_records.extend(source_records)
        self._target_records.extend(target_records)
    
    def _auto_fit_columns(self):
        """Auto-fit column widths based on content."""
        for sheet in [self.sheet1, self.sheet2]:
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column_letter
                
                for cell in column_cells:
                    try:
                        cell_length = len(str(cell.value)) if cell.value else 0
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                
                # Set width with some padding, max 50
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column].width = adjusted_width
    
    def _add_stats_sheet(self, stats: GenerationStats):
        """Add a summary statistics sheet."""
        stats_sheet = self.workbook.create_sheet(title="Generation_Stats")
        
        # Title
        stats_sheet.cell(row=1, column=1, value="Data Generation Summary")
        stats_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        stats_sheet.merge_cells('A1:B1')
        
        # Stats data
        stats_data = stats.to_dict()
        row = 3
        for label, value in stats_data.items():
            stats_sheet.cell(row=row, column=1, value=label)
            stats_sheet.cell(row=row, column=2, value=value)
            stats_sheet.cell(row=row, column=1).font = Font(bold=True)
            row += 1
        
        # Percentages
        row += 1
        stats_sheet.cell(row=row, column=1, value="Distribution Percentages")
        stats_sheet.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1
        
        total_matches = stats.exact_1_to_1_matches + stats.exact_1_to_n_matches
        total_source = stats.source_rows
        
        if total_source > 0:
            percentages = [
                ("Exact Matches", f"{(total_matches / total_source) * 100:.1f}%"),
                ("  - 1:1 Matches", f"{(stats.exact_1_to_1_matches / total_source) * 100:.1f}%"),
                ("  - 1:N Matches", f"{(stats.exact_1_to_n_matches / total_source) * 100:.1f}%"),
                ("Potential Matches", f"{(stats.potential_matches / total_source) * 100:.1f}%"),
                ("Unmatched (Source)", f"{(stats.unmatched_source / total_source) * 100:.1f}%"),
            ]
            
            for label, value in percentages:
                stats_sheet.cell(row=row, column=1, value=label)
                stats_sheet.cell(row=row, column=2, value=value)
                row += 1
        
        # Mapping Keys Section
        row += 2
        stats_sheet.cell(row=row, column=1, value="Mapping Keys for Reconciliation")
        stats_sheet.cell(row=row, column=1).font = Font(bold=True, size=12)
        stats_sheet.merge_cells(f'A{row}:B{row}')
        row += 2
        
        # Get key columns from schemas
        dataset1_keys = [col.name for col in self.scenario.dataset1_schema if col.is_key]
        dataset2_keys = [col.name for col in self.scenario.dataset2_schema if col.is_key]
        dataset1_monetary = [col.name for col in self.scenario.dataset1_schema if col.is_monetary]
        dataset2_monetary = [col.name for col in self.scenario.dataset2_schema if col.is_monetary]
        
        # Dataset 1 keys
        stats_sheet.cell(row=row, column=1, value=f"{self.scenario.dataset1_name}")
        stats_sheet.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        stats_sheet.cell(row=row, column=1, value="  Key Columns:")
        stats_sheet.cell(row=row, column=2, value=", ".join(dataset1_keys))
        row += 1
        stats_sheet.cell(row=row, column=1, value="  Amount Column:")
        stats_sheet.cell(row=row, column=2, value=", ".join(dataset1_monetary))
        row += 2
        
        # Dataset 2 keys
        stats_sheet.cell(row=row, column=1, value=f"{self.scenario.dataset2_name}")
        stats_sheet.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        stats_sheet.cell(row=row, column=1, value="  Key Columns:")
        stats_sheet.cell(row=row, column=2, value=", ".join(dataset2_keys))
        row += 1
        stats_sheet.cell(row=row, column=1, value="  Amount Column:")
        stats_sheet.cell(row=row, column=2, value=", ".join(dataset2_monetary))
        row += 2
        
        # Matching guidance
        stats_sheet.cell(row=row, column=1, value="Recommended Matching Rules:")
        stats_sheet.cell(row=row, column=1).font = Font(bold=True, italic=True)
        row += 1
        stats_sheet.cell(row=row, column=1, value="  1. Match on key columns between datasets")
        row += 1
        stats_sheet.cell(row=row, column=1, value="  2. Compare amount columns for monetary reconciliation")
        row += 1
        stats_sheet.cell(row=row, column=1, value="  3. For 1:N matches, sum target amounts to match source")
        
        # Add example records section
        row += 3
        row = self._add_example_records_section(stats_sheet, stats, row)
        
        # Auto-fit columns
        stats_sheet.column_dimensions['A'].width = 35
        stats_sheet.column_dimensions['B'].width = 40
        stats_sheet.column_dimensions['C'].width = 40
        stats_sheet.column_dimensions['D'].width = 40
        stats_sheet.column_dimensions['E'].width = 40
        stats_sheet.column_dimensions['F'].width = 40
    
    def _add_example_records_section(self, stats_sheet, stats: GenerationStats, row: int) -> int:
        """Add example 1:1 and 1:N records to demonstrate correct generation."""
        # Section header
        stats_sheet.cell(row=row, column=1, value="EXAMPLE RECORDS FOR VERIFICATION")
        stats_sheet.cell(row=row, column=1).font = Font(bold=True, size=14, color="FF0000")
        stats_sheet.merge_cells(f'A{row}:F{row}')
        row += 2
        
        # Get monetary columns for sum calculation
        dataset1_monetary = [col.name for col in self.scenario.dataset1_schema if col.is_monetary]
        dataset2_monetary = [col.name for col in self.scenario.dataset2_schema if col.is_monetary]
        amount_col1 = dataset1_monetary[0] if dataset1_monetary else None
        amount_col2 = dataset2_monetary[0] if dataset2_monetary else None
        
        # 1:1 Example Section
        stats_sheet.cell(row=row, column=1, value="1:1 Match Example (1 Source → 1 Target, amounts must match)")
        stats_sheet.cell(row=row, column=1).font = Font(bold=True, size=12, color="006600")
        stats_sheet.merge_cells(f'A{row}:F{row}')
        row += 1
        
        if stats.example_1_to_1_source and stats.example_1_to_1_target:
            # Source record header
            stats_sheet.cell(row=row, column=1, value=f"Source ({self.scenario.dataset1_name}):")
            stats_sheet.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            
            # Write source record
            source_cols = list(stats.example_1_to_1_source.keys())
            for col_idx, col_name in enumerate(source_cols, 1):
                cell = stats_sheet.cell(row=row, column=col_idx, value=col_name)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
            row += 1
            for col_idx, col_name in enumerate(source_cols, 1):
                value = stats.example_1_to_1_source.get(col_name)
                stats_sheet.cell(row=row, column=col_idx, value=value)
            row += 2
            
            # Target record header
            stats_sheet.cell(row=row, column=1, value=f"Target ({self.scenario.dataset2_name}):")
            stats_sheet.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            
            target_cols = list(stats.example_1_to_1_target.keys())
            for col_idx, col_name in enumerate(target_cols, 1):
                cell = stats_sheet.cell(row=row, column=col_idx, value=col_name)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
            row += 1
            for col_idx, col_name in enumerate(target_cols, 1):
                value = stats.example_1_to_1_target.get(col_name)
                stats_sheet.cell(row=row, column=col_idx, value=value)
            row += 2
            
            # Verification
            source_amt = stats.example_1_to_1_source.get(amount_col1, 0) if amount_col1 else 0
            target_amt = stats.example_1_to_1_target.get(amount_col2, 0) if amount_col2 else 0
            stats_sheet.cell(row=row, column=1, value="✓ VERIFICATION:")
            stats_sheet.cell(row=row, column=1).font = Font(bold=True, color="006600")
            stats_sheet.cell(row=row, column=2, value=f"Source Amount: {source_amt:.2f}")
            stats_sheet.cell(row=row, column=3, value=f"Target Amount: {target_amt:.2f}")
            stats_sheet.cell(row=row, column=4, value=f"Match: {'YES' if abs(source_amt - target_amt) < 0.01 else 'NO'}")
            row += 1
        else:
            stats_sheet.cell(row=row, column=1, value="No 1:1 examples generated (0% configured)")
            row += 1
        
        row += 2
        
        # 1:N Example Section
        stats_sheet.cell(row=row, column=1, value="1:N Match Example (1 Source → N Targets, amounts must SUM to source)")
        stats_sheet.cell(row=row, column=1).font = Font(bold=True, size=12, color="0000FF")
        stats_sheet.merge_cells(f'A{row}:F{row}')
        row += 1
        
        if stats.example_1_to_n_source and stats.example_1_to_n_targets:
            # Source record header
            stats_sheet.cell(row=row, column=1, value=f"Source ({self.scenario.dataset1_name}):")
            stats_sheet.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            
            # Write source record
            source_cols = list(stats.example_1_to_n_source.keys())
            for col_idx, col_name in enumerate(source_cols, 1):
                cell = stats_sheet.cell(row=row, column=col_idx, value=col_name)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
            row += 1
            for col_idx, col_name in enumerate(source_cols, 1):
                value = stats.example_1_to_n_source.get(col_name)
                stats_sheet.cell(row=row, column=col_idx, value=value)
            row += 2
            
            # Target records header
            n_targets = len(stats.example_1_to_n_targets)
            stats_sheet.cell(row=row, column=1, value=f"Targets ({self.scenario.dataset2_name}) - {n_targets} records:")
            stats_sheet.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            
            target_cols = list(stats.example_1_to_n_targets[0].keys())
            for col_idx, col_name in enumerate(target_cols, 1):
                cell = stats_sheet.cell(row=row, column=col_idx, value=col_name)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
            row += 1
            
            # Write each target record
            target_sum = 0
            for idx, target in enumerate(stats.example_1_to_n_targets):
                for col_idx, col_name in enumerate(target_cols, 1):
                    value = target.get(col_name)
                    stats_sheet.cell(row=row, column=col_idx, value=value)
                if amount_col2:
                    target_sum += target.get(amount_col2, 0)
                row += 1
            
            row += 1
            
            # Verification with sum calculation
            source_amt = stats.example_1_to_n_source.get(amount_col1, 0) if amount_col1 else 0
            stats_sheet.cell(row=row, column=1, value="✓ VERIFICATION:")
            stats_sheet.cell(row=row, column=1).font = Font(bold=True, color="0000FF")
            stats_sheet.cell(row=row, column=2, value=f"Source Amount: {source_amt:.2f}")
            stats_sheet.cell(row=row, column=3, value=f"Target SUM: {target_sum:.2f}")
            match_status = 'YES' if abs(source_amt - target_sum) < 0.01 else 'NO'
            stats_sheet.cell(row=row, column=4, value=f"SUM Match: {match_status}")
            row += 1
            
            # Breakdown
            row += 1
            stats_sheet.cell(row=row, column=1, value="Amount Breakdown:")
            stats_sheet.cell(row=row, column=1).font = Font(italic=True)
            row += 1
            for idx, target in enumerate(stats.example_1_to_n_targets):
                amt = target.get(amount_col2, 0) if amount_col2 else 0
                stats_sheet.cell(row=row, column=1, value=f"  Target {idx + 1}:")
                stats_sheet.cell(row=row, column=2, value=f"{amt:.2f}")
                row += 1
            stats_sheet.cell(row=row, column=1, value="  TOTAL:")
            stats_sheet.cell(row=row, column=1).font = Font(bold=True)
            stats_sheet.cell(row=row, column=2, value=f"{target_sum:.2f}")
            stats_sheet.cell(row=row, column=2).font = Font(bold=True)
            row += 1
        else:
            stats_sheet.cell(row=row, column=1, value="No 1:N examples generated (0% 1:N ratio configured)")
            row += 1
        
        row += 2
        
        # Partial Match Example Section (per MS Copilot Finance docs)
        stats_sheet.cell(row=row, column=1, value="PARTIAL MATCH Example (Substring Matching per MS Copilot Finance)")
        stats_sheet.cell(row=row, column=1).font = Font(bold=True, size=12, color="800080")  # Purple
        stats_sheet.merge_cells(f'A{row}:F{row}')
        row += 1
        
        # Get key columns - use the scenario's primary/secondary key column properties
        # These are the actual mapping key columns, not all key columns
        key_col1 = self.scenario.primary_key_column
        key_col2 = self.scenario.secondary_key_column
        
        if stats.example_partial_match_source and stats.example_partial_match_target:
            # Explanation
            stats_sheet.cell(row=row, column=1, value="Per MS Docs: Partial match = substring match on key + equal amounts → Potentially Matched")
            stats_sheet.cell(row=row, column=1).font = Font(italic=True)
            stats_sheet.merge_cells(f'A{row}:F{row}')
            row += 2
            
            # Source record
            stats_sheet.cell(row=row, column=1, value=f"Source ({self.scenario.dataset1_name}):")
            stats_sheet.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            
            source_cols = list(stats.example_partial_match_source.keys())
            for col_idx, col_name in enumerate(source_cols, 1):
                cell = stats_sheet.cell(row=row, column=col_idx, value=col_name)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
            row += 1
            for col_idx, col_name in enumerate(source_cols, 1):
                value = stats.example_partial_match_source.get(col_name)
                stats_sheet.cell(row=row, column=col_idx, value=value)
            row += 2
            
            # Target record
            stats_sheet.cell(row=row, column=1, value=f"Target ({self.scenario.dataset2_name}):")
            stats_sheet.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            
            target_cols = list(stats.example_partial_match_target.keys())
            for col_idx, col_name in enumerate(target_cols, 1):
                cell = stats_sheet.cell(row=row, column=col_idx, value=col_name)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
            row += 1
            for col_idx, col_name in enumerate(target_cols, 1):
                value = stats.example_partial_match_target.get(col_name)
                stats_sheet.cell(row=row, column=col_idx, value=value)
            row += 2
            
            # Verification - show the substring relationship
            source_ref = stats.example_partial_match_source.get(key_col1, "") if key_col1 else ""
            target_ref = stats.example_partial_match_target.get(key_col2, "") if key_col2 else ""
            source_amt = stats.example_partial_match_source.get(amount_col1, 0) if amount_col1 else 0
            target_amt = stats.example_partial_match_target.get(amount_col2, 0) if amount_col2 else 0
            
            # Check substring relationship
            is_superset = str(source_ref) in str(target_ref)
            is_subset = str(target_ref) in str(source_ref)
            substring_match = is_superset or is_subset
            substring_type = "Target contains Source" if is_superset else ("Source contains Target" if is_subset else "N/A")
            
            stats_sheet.cell(row=row, column=1, value="✓ VERIFICATION:")
            stats_sheet.cell(row=row, column=1).font = Font(bold=True, color="800080")
            row += 1
            stats_sheet.cell(row=row, column=1, value=f"  Source Key:")
            stats_sheet.cell(row=row, column=2, value=str(source_ref))
            row += 1
            stats_sheet.cell(row=row, column=1, value=f"  Target Key:")
            stats_sheet.cell(row=row, column=2, value=str(target_ref))
            row += 1
            stats_sheet.cell(row=row, column=1, value=f"  Substring Match:")
            stats_sheet.cell(row=row, column=2, value=f"{'YES' if substring_match else 'NO'} ({substring_type})")
            row += 1
            stats_sheet.cell(row=row, column=1, value=f"  Amount Equal:")
            stats_sheet.cell(row=row, column=2, value=f"{'YES' if abs(source_amt - target_amt) < 0.01 else 'NO'} (Source: {source_amt:.2f}, Target: {target_amt:.2f})")
            row += 1
            stats_sheet.cell(row=row, column=1, value=f"  Classification:")
            stats_sheet.cell(row=row, column=1).font = Font(bold=True)
            stats_sheet.cell(row=row, column=2, value="POTENTIALLY MATCHED (per MS Copilot Finance)")
            stats_sheet.cell(row=row, column=2).font = Font(bold=True, color="800080")
            row += 1
        else:
            stats_sheet.cell(row=row, column=1, value="No partial match examples generated in this run")
            row += 1
        
        return row
    
    # ------------------------------------------------------------------ #
    #  CSV helpers                                                        #
    # ------------------------------------------------------------------ #

    def _write_csv(self, path: Path, columns: List[str], records: List[dict]):
        """Write a list of record dicts to a CSV file."""
        with open(path, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.DictWriter(fh, fieldnames=columns, extrasaction="ignore")
            writer.writeheader()
            for record in records:
                # Ensure date values are strings in CSV too
                row = {k: self._format_value(v) for k, v in record.items()}
                writer.writerow(row)

    # ------------------------------------------------------------------ #
    #  Finalise & close                                                  #
    # ------------------------------------------------------------------ #

    def finalize(self, stats: GenerationStats) -> dict:
        """Finalize and save the workbook + CSVs into the output subfolder.

        Returns a dict with the paths of the generated files.
        """
        self._auto_fit_columns()
        self._add_stats_sheet(stats)
        
        # --- XLSX -------------------------------------------------------- #
        xlsx_name = self.output_dir.name + ".xlsx"
        xlsx_path = self.output_dir / xlsx_name
        self.workbook.save(str(xlsx_path))

        # --- CSVs -------------------------------------------------------- #
        columns1 = self.scenario.get_dataset1_columns()
        columns2 = self.scenario.get_dataset2_columns()

        # Sanitise sheet names for use as file names
        csv1_name = self.scenario.dataset1_name.replace(" ", "_") + ".csv"
        csv2_name = self.scenario.dataset2_name.replace(" ", "_") + ".csv"

        csv1_path = self.output_dir / csv1_name
        csv2_path = self.output_dir / csv2_name

        self._write_csv(csv1_path, columns1, self._source_records)
        self._write_csv(csv2_path, columns2, self._target_records)

        return {
            "dir": str(self.output_dir.absolute()),
            "xlsx": str(xlsx_path.absolute()),
            "csv_source": str(csv1_path.absolute()),
            "csv_target": str(csv2_path.absolute()),
        }
    
    def close(self):
        """Close the workbook."""
        self.workbook.close()
